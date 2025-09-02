import tkinter as tk
from tkinter import Canvas, Scrollbar, Menu
import pandas as pd
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.tooltip import ToolTip
from tkinter import filedialog, Text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import os
from datetime import datetime

class ExcelComparatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Comparison Tool")
        self.root.geometry("850x900")

        self.df1 = None
        self.df2 = None
        self.df1_original_cols = None
        self.df2_original_cols = None
        self.common_cols_list = []
        self.column_vars = {}
        self.src_to_tgt_case_map = {}

        self.setup_gui()

    def normalize_for_comparison(self, series, is_case_insensitive_data, should_trim_whitespace):
        s = series.copy()
        s = s.astype(str)
        s = s.str.replace(r'\.0$', '', regex=True)
        s_lower_for_nulls = s.str.lower().str.strip()
        s[s_lower_for_nulls.isin(['nan', '<na>', 'none', 'nat'])] = ''
        if should_trim_whitespace:
            s = s.str.strip()
            s = s.str.replace(r'\s+', ' ', regex=True)
        if is_case_insensitive_data:
            s = s.str.lower()
        return s

    def setup_gui(self):
        menubar = Menu(self.root)
        self.root.config(menu=menubar)
        theme_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Theme", menu=theme_menu)
        theme_menu.add_radiobutton(label="Litera (Light)", command=lambda: self.root.style.theme_use('litera'))
        theme_menu.add_radiobutton(label="Superhero (Dark)", command=lambda: self.root.style.theme_use('superhero'))

        inputs = ttk.Frame(self.root, padding=10)
        inputs.pack(fill=X, padx=10, pady=5)
        inputs.columnconfigure(1, weight=1)

        self.source_rows_var = tk.StringVar(value="Row Count: -")
        self.target_rows_var = tk.StringVar(value="Row Count: -")

        ttk.Label(inputs, text="Source File:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.source_path_entry = ttk.Entry(inputs, width=70)
        self.source_path_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        browse_src_btn = ttk.Button(inputs, text="Browse", style="outline.TButton", command=lambda: [self.source_path_entry.delete(0, 'end'), self.source_path_entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")]))])
        browse_src_btn.grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(inputs, text="Source Header Row:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.source_header_entry = ttk.Entry(inputs, width=10)
        self.source_header_entry.insert(0, "1")
        self.source_header_entry.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(inputs, textvariable=self.source_rows_var, style="secondary.TLabel").grid(row=1, column=2, sticky='w', padx=10)

        ttk.Label(inputs, text="Target File:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.target_path_entry = ttk.Entry(inputs, width=70)
        self.target_path_entry.grid(row=2, column=1, sticky='ew', padx=5, pady=5)
        browse_tgt_btn = ttk.Button(inputs, text="Browse", style="outline.TButton", command=lambda: [self.target_path_entry.delete(0, 'end'), self.target_path_entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")]))])
        browse_tgt_btn.grid(row=2, column=2, padx=5, pady=5)
        
        ttk.Label(inputs, text="Target Header Row:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.target_header_entry = ttk.Entry(inputs, width=10)
        self.target_header_entry.insert(0, "1")
        self.target_header_entry.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        ttk.Label(inputs, textvariable=self.target_rows_var, style="secondary.TLabel").grid(row=3, column=2, sticky='w', padx=10)
        
        ttk.Label(inputs, text="Output File Name:").grid(row=4, column=0, sticky='w', padx=5, pady=5)
        self.output_file_entry = ttk.Entry(inputs, width=50)
        self.output_file_entry.insert(0, "comparison_output")
        self.output_file_entry.grid(row=4, column=1, sticky='w', padx=5, pady=5)
        
        options_frame = ttk.Labelframe(self.root, text="Comparison Options", padding=10)
        options_frame.pack(fill=X, padx=10, pady=5)
        
        self.case_insensitive_cols_var = tk.BooleanVar(value=True)
        ci_cols_check = ttk.Checkbutton(options_frame, text="Case-Insensitive Column Names", variable=self.case_insensitive_cols_var)
        ci_cols_check.pack(side=LEFT, padx=10)
        ToolTip(ci_cols_check, "If checked, 'ColumnA' and 'columna' are treated as the same column.")
        
        self.case_insensitive_data_var = tk.BooleanVar(value=True)
        ci_data_check = ttk.Checkbutton(options_frame, text="Case-Insensitive Data", variable=self.case_insensitive_data_var)
        ci_data_check.pack(side=LEFT, padx=10)
        ToolTip(ci_data_check, "If checked, 'Apple' and 'apple' are treated as the same value during comparison.")
        
        self.trim_whitespace_var = tk.BooleanVar(value=True)
        trim_check = ttk.Checkbutton(options_frame, text="Trim Whitespace", variable=self.trim_whitespace_var)
        trim_check.pack(side=LEFT, padx=10)
        ToolTip(trim_check, "If checked, removes leading/trailing spaces and reduces multiple internal spaces to one.")

        load_cols_btn = ttk.Button(self.root, text="Load Columns to Select", command=self.load_and_display_columns, style="primary.TButton")
        load_cols_btn.pack(pady=(5,0))
        
        col_select_frame = ttk.Labelframe(self.root, text="Select Columns for Comparison (in Source File Order)", padding=10)
        col_select_frame.pack(fill=X, padx=10, pady=5)
        canvas = Canvas(col_select_frame, height=120, borderwidth=0)
        scrollbar = ttk.Scrollbar(col_select_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=LEFT, fill="both", expand=True)
        scrollbar.pack(side=RIGHT, fill="y")
        
        sheet_select = ttk.Labelframe(self.root, text="Sheets to Generate", padding=10)
        sheet_select.pack(fill=X, padx=10, pady=5)
        self.col_sheet_var = tk.BooleanVar(value=True); self.row_sheet_var = tk.BooleanVar(value=True)
        self.uniq_sheet_var = tk.BooleanVar(value=True); self.stats_sheet_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(sheet_select, text="Column Names", variable=self.col_sheet_var).pack(side=LEFT, padx=10)
        ttk.Checkbutton(sheet_select, text="Row Comparison", variable=self.row_sheet_var).pack(side=LEFT, padx=10)
        ttk.Checkbutton(sheet_select, text="Unique Values", variable=self.uniq_sheet_var).pack(side=LEFT, padx=10)
        ttk.Checkbutton(sheet_select, text="Summary Stats", variable=self.stats_sheet_var).pack(side=LEFT, padx=10)
        
        button_frame = ttk.Frame(self.root)
        button_frame.pack(pady=10)
        compare_btn = ttk.Button(button_frame, text="Compare Files", command=self.compare_files, style="success.TButton")
        compare_btn.pack(side=LEFT, padx=10)
        clear_btn = ttk.Button(button_frame, text="Clear", command=self.clear_fields, style="danger.outline.TButton")
        clear_btn.pack(side=LEFT, padx=10)

        log_frame = ttk.Labelframe(self.root, text="Log Output", padding=10)
        log_frame.pack(pady=10, padx=10, fill=BOTH, expand=True)
        self.log_output = Text(log_frame, height=10, width=100, wrap='word', relief="solid", borderwidth=1, font=("Consolas", 10))
        self.log_output.pack(fill=BOTH, expand=True)

    def log(self, message):
        self.log_output.insert(END, message)
        self.root.update_idletasks()
        
    def load_and_display_columns(self):
        self.log("Loading files to find common columns...\n")
        try:
            for widget in self.scrollable_frame.winfo_children():
                widget.destroy()
            self.df1, self.df2, self.column_vars = None, None, {}
            self.src_to_tgt_case_map = {}

            src_path = self.source_path_entry.get(); tgt_path = self.target_path_entry.get()
            if not src_path or not tgt_path:
                self.log("❌ Error: Please select both source and target files first.\n"); return

            src_header = int(self.source_header_entry.get()) - 1
            tgt_header = int(self.target_header_entry.get()) - 1

            self.df1 = pd.read_excel(src_path, header=src_header)
            self.df2 = pd.read_excel(tgt_path, header=tgt_header)
            self.source_rows_var.set(f"Row Count: {len(self.df1)}")
            self.target_rows_var.set(f"Row Count: {len(self.df2)}")
            
            self.df1_original_cols = self.df1.columns.copy()
            self.df2_original_cols = self.df2.columns.copy()
            
            if self.case_insensitive_cols_var.get():
                src_lower_to_orig = {col.lower(): col for col in self.df1.columns}
                tgt_lower_to_orig = {col.lower(): col for col in self.df2.columns}
                common_cols_lower = set(src_lower_to_orig.keys()).intersection(tgt_lower_to_orig.keys())
                self.common_cols_list = [src_lower_to_orig[c.lower()] for c in self.df1.columns if c.lower() in common_cols_lower]
                self.src_to_tgt_case_map = {src_lower_to_orig[c_lower]: tgt_lower_to_orig[c_lower] for c_lower in common_cols_lower}
            else:
                df2_cols = set(self.df2.columns)
                self.common_cols_list = [col for col in self.df1.columns if col in df2_cols]
                self.src_to_tgt_case_map = {col: col for col in self.common_cols_list}

            if not self.common_cols_list:
                self.log("⚠️ Warning: No common columns found.\n"); return

            max_cols = 4
            for i, col_name in enumerate(self.common_cols_list):
                var = tk.BooleanVar(value=True)
                self.column_vars[col_name] = var
                row, col = divmod(i, max_cols)
                cb = ttk.Checkbutton(self.scrollable_frame, text=col_name, variable=var)
                cb.grid(row=row, column=col, sticky='w', padx=5, pady=2)

            self.log(f"✅ Found {len(self.common_cols_list)} common columns. Please make your selections.\n")

        except Exception as e:
            self.log(f"❌ Error loading columns: {str(e)}\n")

    def compare_files(self):
        self.log_output.delete("1.0", "end")
        self.log("Starting comparison...\n")

        if self.df1 is None or self.df2 is None:
            self.log("❌ Error: Please use 'Load Columns to Select' before comparing.\n"); return

        selected_cols_src = [col for col, var in self.column_vars.items() if var.get()]
        if not selected_cols_src:
            self.log("❌ Error: No columns selected for comparison.\n"); return
        
        base_filename = self.output_file_entry.get()
        if not base_filename:
            self.log("❌ Error: Please provide an output file name.\n"); return

        try:
            name, ext = os.path.splitext(base_filename); ext = ext or ".xlsx"
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            output_file = os.path.join(os.getcwd(), f"{name}_{timestamp}{ext}")

            wb = Workbook(); wb.remove(wb.active)
            bold_font = Font(bold=True)
            
            is_case_insensitive_data = self.case_insensitive_data_var.get()
            should_trim = self.trim_whitespace_var.get()

            # --- Sheet 1: Column Name Comparison (FIXED) ---
            if self.col_sheet_var.get():
                self.log("Generating 'Column Name Comparison' sheet...\n")
                sheet1 = wb.create_sheet("Column Name Comparison")
                sheet1.append(["Column Name", "In Source File", "In Target File"])
                
                if self.case_insensitive_cols_var.get():
                    lower_to_orig_casing = {}
                    col_status = {}
                    
                    for col in self.df1_original_cols:
                        lower_col = col.lower()
                        if lower_col not in lower_to_orig_casing:
                            lower_to_orig_casing[lower_col] = col
                        if lower_col not in col_status:
                            col_status[lower_col] = {'in_src': False, 'in_tgt': False}
                        col_status[lower_col]['in_src'] = True
                        
                    for col in self.df2_original_cols:
                        lower_col = col.lower()
                        if lower_col not in lower_to_orig_casing:
                            lower_to_orig_casing[lower_col] = col
                        if lower_col not in col_status:
                            col_status[lower_col] = {'in_src': False, 'in_tgt': False}
                        col_status[lower_col]['in_tgt'] = True
                    
                    # Display in source file order
                    display_order = [key for key in lower_to_orig_casing if key in col_status]
                    for lower_col in display_order:
                        display_name = lower_to_orig_casing[lower_col]
                        status = col_status[lower_col]
                        sheet1.append([display_name, 'Yes' if status['in_src'] else 'No', 'Yes' if status['in_tgt'] else 'No'])

                else: # Case-sensitive
                    all_cols_ordered, seen_cols = [], set()
                    for col in list(self.df1_original_cols) + list(self.df2_original_cols):
                        if col not in seen_cols:
                            all_cols_ordered.append(col); seen_cols.add(col)
                    for col in all_cols_ordered:
                        sheet1.append([col, "Yes" if col in self.df1_original_cols else "No", "Yes" if col in self.df2_original_cols else "No"])

            if self.row_sheet_var.get():
                self.log("Generating 'Row Comparison' sheet...\n")
                selected_cols_tgt = [self.src_to_tgt_case_map.get(col, col) for col in selected_cols_src]
                df1_compare = self.df1[selected_cols_src].copy().reset_index()
                df2_compare = self.df2[selected_cols_tgt].copy().reset_index()
                df2_compare.rename(columns=dict(zip(selected_cols_tgt, selected_cols_src)), inplace=True)

                self.log("  ▶ Normalizing data for accurate matching...\n")
                for col in selected_cols_src:
                    df1_compare[col] = self.normalize_for_comparison(df1_compare[col], is_case_insensitive_data, should_trim)
                    df2_compare[col] = self.normalize_for_comparison(df2_compare[col], is_case_insensitive_data, should_trim)

                df1_compare.drop_duplicates(subset=selected_cols_src, inplace=True)
                df2_compare.drop_duplicates(subset=selected_cols_src, inplace=True)

                merged_df = pd.merge(df1_compare, df2_compare, on=selected_cols_src, how='outer', 
                                     indicator='indicator_col', suffixes=('_src', '_tgt'))
                
                output_rows = []
                for row in merged_df.itertuples(index=False):
                    status, original_row_data = "", []
                    if row.indicator_col == 'left_only':
                        status = "Only in Source"
                        original_row_data = self.df1[selected_cols_src].loc[int(row.index_src)].tolist()
                    elif row.indicator_col == 'right_only':
                        status = "Only in Target"
                        original_row_data = self.df2[selected_cols_tgt].loc[int(row.index_tgt)].tolist()
                    elif row.indicator_col == 'both':
                        status = "In Both Files"
                        original_row_data = self.df1[selected_cols_src].loc[int(row.index_src)].tolist()
                    output_rows.append([status] + original_row_data)

                sheet2 = wb.create_sheet("Row Comparison")
                sheet2.append(['Status'] + selected_cols_src)
                for cell in sheet2[1]: cell.font = bold_font
                for row_data in output_rows:
                    sheet2.append([str(item) if pd.isna(item) else item for item in row_data])

            # --- RESTORED: Unique Values and Summary Stats ---
            if self.uniq_sheet_var.get():
                self.log("Generating 'Unique Values' sheet based on selection...\n")
                sheet3 = wb.create_sheet("Unique Values")
                unique_data = {}; max_unique_rows = 0
                for col in selected_cols_src:
                    src_unique = set(self.df1[col].dropna().astype(str).unique())
                    tgt_col = self.src_to_tgt_case_map.get(col, col)
                    tgt_unique = set(self.df2[tgt_col].dropna().astype(str).unique())
                    only_src = sorted(list(src_unique - tgt_unique))
                    only_tgt = sorted(list(tgt_unique - src_unique))
                    unique_data[col] = {'source': only_src, 'target': only_tgt}
                    max_unique_rows = max(max_unique_rows, len(only_src), len(only_tgt))
                
                col_idx = 1
                for col in selected_cols_src:
                    sheet3.cell(row=1, column=col_idx, value=col).font = bold_font
                    sheet3.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
                    sheet3.cell(row=2, column=col_idx, value="Only in Source").font = bold_font
                    sheet3.cell(row=2, column=col_idx + 1, value="Only in Target").font = bold_font
                    col_idx += 2
                
                for i in range(max_unique_rows):
                    col_idx = 1
                    for col in selected_cols_src:
                        if i < len(unique_data[col]['source']): sheet3.cell(row=i + 3, column=col_idx, value=unique_data[col]['source'][i])
                        if i < len(unique_data[col]['target']): sheet3.cell(row=i + 3, column=col_idx + 1, value=unique_data[col]['target'][i])
                        col_idx += 2
            
            if self.stats_sheet_var.get():
                self.log("Generating 'Summary Stats' sheet...\n")
                sheet4 = wb.create_sheet("Summary Stats")
                
                numeric_cols = [col for col in selected_cols_src if pd.api.types.is_numeric_dtype(self.df1[col]) and pd.api.types.is_numeric_dtype(self.df2[self.src_to_tgt_case_map.get(col, col)])]

                if not numeric_cols:
                    self.log("  ▶ No common numeric columns found in selection for stats.\n")
                    sheet4.cell(row=1, column=1, value="No common numeric columns found among selected columns.")
                else:
                    stats_to_calc = ['count', 'mean', 'median', 'min', 'max', 'sum']
                    header_alignment = Alignment(horizontal='center', vertical='center')
                    col_idx = 1
                    for col in numeric_cols:
                        cell = sheet4.cell(row=1, column=col_idx, value=col)
                        cell.alignment = header_alignment; cell.font = bold_font
                        sheet4.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
                        sheet4.cell(row=2, column=col_idx, value="Stat").font = bold_font
                        sheet4.cell(row=2, column=col_idx + 1, value="Source").font = bold_font
                        sheet4.cell(row=2, column=col_idx + 2, value="Target").font = bold_font
                        col_idx += 3
                    
                    row_idx = 3
                    for stat in stats_to_calc:
                        col_idx = 1
                        for col in numeric_cols:
                            try:
                                tgt_col = self.src_to_tgt_case_map.get(col, col)
                                s_val = getattr(self.df1[col], stat)(); t_val = getattr(self.df2[tgt_col], stat)()
                            except Exception:
                                s_val, t_val = 'N/A', 'N/A'
                            sheet4.cell(row=row_idx, column=col_idx, value=stat)
                            sheet4.cell(row=row_idx, column=col_idx + 1, value=s_val)
                            sheet4.cell(row=row_idx, column=col_idx + 2, value=t_val)
                            col_idx += 3
                        row_idx += 1

            if not wb.sheetnames:
                 self.log("⚠️ Warning: No sheets selected. Nothing to save.\n"); return
            
            self.log(f"Saving file to {output_file}...\n")
            wb.save(output_file)
            self.log(f"\n✅ Comparison successful!\nOutput saved to: {output_file}\n")

        except Exception as e:
            self.log(f"❌ An unexpected error occurred during comparison: {str(e)}\n")

    def clear_fields(self):
        self.source_path_entry.delete(0, END); self.target_path_entry.delete(0, END)
        self.output_file_entry.delete(0, END); self.output_file_entry.insert(0, "comparison_output")
        self.source_header_entry.delete(0, END); self.source_header_entry.insert(0, "1")
        self.target_header_entry.delete(0, END); self.target_header_entry.insert(0, "1")
        self.case_insensitive_cols_var.set(True); self.case_insensitive_data_var.set(True)
        self.trim_whitespace_var.set(True)
        self.source_rows_var.set("Row Count: -"); self.target_rows_var.set("Row Count: -")
        self.df1, self.df2 = None, None
        self.column_vars = {}; self.src_to_tgt_case_map = {}
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.log_output.delete("1.0", "end")
        self.log("Fields cleared. Ready for new comparison.\n")

if __name__ == "__main__":
    root = ttk.Window(themename="litera")
    app = ExcelComparatorApp(root)
    root.mainloop()